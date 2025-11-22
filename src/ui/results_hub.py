import json
import os
import sqlite3
import sys
from pathlib import Path
from typing import Any

from PyQt6.QtCore import QSize, Qt, QUrl
from PyQt6.QtGui import QIcon, QPixmap
from PyQt6.QtNetwork import QNetworkAccessManager, QNetworkRequest
from PyQt6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QCheckBox,
    QComboBox,
    QFileDialog,
    QFormLayout,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QSplitter,
    QStackedWidget,
    QStyledItemDelegate,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

# Add project root to path for imports
current_dir = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(current_dir))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

DB_PATH = Path(PROJECT_ROOT) / "data" / "databases" / "products.db"


class LargeTextDelegate(QStyledItemDelegate):
    """Custom delegate that provides a larger editor for table cells."""

    def createEditor(self, parent, option, index):
        """Create a larger line edit for better text visibility."""
        editor = QLineEdit(parent)
        editor.setMinimumHeight(40)  # Make editor taller
        editor.setStyleSheet("""
            QLineEdit {
                font-size: 14px;
                padding: 5px;
            }
        """)
        return editor


class ConsolidationWidget(QWidget):
    """Widget for manually consolidating product data from multiple sources."""

    def __init__(self, sku_data: dict, results_hub_parent):
        """
        Args:
            sku_data: Dict with structure:
                {
                    "sku": "12345",
                    "scrapers": {
                        "central_pet": {"Name": "...", "Brand": "...", ...},
                        "petfood": {"Name": "...", "Brand": "...", ...}
                    }
                }
            results_hub_parent: Reference to the ResultsHub instance
        """
        super().__init__()
        self.sku_data = sku_data
        self.sku = sku_data["sku"]
        self.scrapers = list(sku_data["scrapers"].keys())
        self.results_hub = results_hub_parent

        # Consolidation queue from parent
        self.consolidation_queue = getattr(results_hub_parent, "consolidation_queue", [])
        self.current_index = getattr(results_hub_parent, "consolidation_index", 0)

        # Store user selections: {field_name: scraper_name}
        self.selections = {}

        # Load previous selections if this product was already consolidated
        self._load_previous_selections()

        # Fields to consolidate (only scraped fields, not Price or SKU)
        # Order: Name, Brand, Weight, Images
        self.fields = ["Name", "Brand", "Weight", "Images"]

        self.create_ui()

    def create_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Initialize thumbnail layout early to avoid AttributeError
        self.thumbnail_container = QWidget()
        self.thumbnail_layout = QHBoxLayout(self.thumbnail_container)
        self.thumbnail_layout.setSpacing(10)
        self.thumbnail_layout.setContentsMargins(5, 5, 5, 5)
        self.thumbnail_layout.addStretch()

        # Title with progress
        if self.consolidation_queue:
            progress_text = f"Product {self.current_index + 1} of {len(self.consolidation_queue)}"
            title = QLabel(f"Consolidating SKU: {self.sku} ({progress_text})")
        else:
            title = QLabel(f"Consolidating SKU: {self.sku}")
        title.setStyleSheet("font-size: 18px; font-weight: bold;")
        layout.addWidget(title)

        # Main content: Two-column layout (Fields | Images)
        content_layout = QHBoxLayout()
        content_layout.setSpacing(30)

        # === LEFT COLUMN: Product Details ===
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(0, 0, 0, 0)

        # Instructions
        instructions = QLabel("Select the best data source for each field:")
        instructions.setStyleSheet("color: #aaa; font-size: 12px; margin-bottom: 10px;")
        left_layout.addWidget(instructions)

        # Scroll area for fields (Name, Brand, Weight only - Images handled separately)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll.setMinimumWidth(400)

        form_widget = QWidget()
        form_layout = QFormLayout(form_widget)
        form_layout.setSpacing(20)
        form_layout.setContentsMargins(10, 10, 10, 10)

        self.field_widgets = {}

        # Only show Name, Brand, Weight in left column
        for field in ["Name", "Brand", "Weight"]:
            field_group = self._create_field_selector(field)
            form_layout.addRow(field_group)

        scroll.setWidget(form_widget)
        left_layout.addWidget(scroll)
        left_layout.addStretch()

        content_layout.addWidget(left_widget)

        # === RIGHT COLUMN: Images ===
        right_widget = QWidget()
        right_widget.setMinimumWidth(500)  # Set on widget, not layout
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(0, 0, 0, 0)

        # Image source selector
        image_selector_label = QLabel("**Images:**")
        image_selector_label.setStyleSheet("font-weight: bold; font-size: 14px;")
        right_layout.addWidget(image_selector_label)

        # Dropdown to select image source
        self.image_source_combo = QComboBox()
        self.image_source_combo.addItem("-- Select Source --", None)

        # Collect all image sources
        self.image_sources = {}  # {scraper_name: [image_urls]}
        for scraper in self.scrapers:
            scraper_data = self.sku_data["scrapers"][scraper]
            images = scraper_data.get("Images", [])
            if images and isinstance(images, list) and len(images) > 0:
                self.image_sources[scraper] = images
                self.image_source_combo.addItem(f"{scraper} ({len(images)} images)", scraper)

        self.image_source_combo.currentIndexChanged.connect(self.on_image_source_changed)
        right_layout.addWidget(self.image_source_combo)

        # Pre-select moved to after thumbnail_layout init to avoid AttributeError

        # Main image preview
        self.main_image_label = QLabel("Select an image source to preview")
        self.main_image_label.setMinimumSize(400, 400)
        self.main_image_label.setMaximumSize(500, 500)
        self.main_image_label.setStyleSheet("""
            QLabel {
                border: 2px solid #555;
                background-color: #1a1a1a;
                padding: 10px;
            }
        """)
        self.main_image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.main_image_label.setScaledContents(False)
        right_layout.addWidget(self.main_image_label)

        # Scrollable thumbnail container (no label)
        thumbnail_scroll = QScrollArea()
        thumbnail_scroll.setWidgetResizable(True)
        thumbnail_scroll.setFixedHeight(120)
        thumbnail_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        thumbnail_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

        # thumbnail_container initialized at top of create_ui

        thumbnail_scroll.setWidget(self.thumbnail_container)
        right_layout.addWidget(thumbnail_scroll)

        # Pre-select if we have a previous image selection (Now safe to trigger signal)
        if "Images" in self.selections:
            saved_scraper = self.selections["Images"]
            for i in range(self.image_source_combo.count()):
                if self.image_source_combo.itemData(i) == saved_scraper:
                    self.image_source_combo.setCurrentIndex(i)
                    break
        
        right_layout.addStretch()

        content_layout.addWidget(right_widget)
        layout.addLayout(content_layout)

        # Action buttons at bottom
        button_layout = QHBoxLayout()

        # Navigation buttons (if in queue mode)
        # Navigation buttons (if in queue mode)
        is_queue_mode = self.consolidation_queue and len(self.consolidation_queue) > 1
        
        if is_queue_mode:
            prev_btn = QPushButton("← Previous")
            prev_btn.setProperty("class", "secondary")
            prev_btn.clicked.connect(self.go_to_previous)
            if self.current_index == 0:
                prev_btn.setEnabled(False)
            button_layout.addWidget(prev_btn)

            next_btn = QPushButton("Next →")
            next_btn.setProperty("class", "primary")
            next_btn.clicked.connect(self.go_to_next)
            if self.current_index >= len(self.consolidation_queue) - 1:
                next_btn.setText("Finish")
            button_layout.addWidget(next_btn)

        button_layout.addStretch()

        cancel_btn = QPushButton("❌ Cancel")
        cancel_btn.setProperty("class", "secondary")
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)

        # Only show explicit Save button in single mode (Queue mode uses Next/Finish)
        if not is_queue_mode:
            save_btn = QPushButton("💾 Save && Close")
            save_btn.setProperty("class", "success")
            save_btn.clicked.connect(self.save_and_close)
            button_layout.addWidget(save_btn)

        layout.addLayout(button_layout)

    def _create_field_selector(self, field_name: str) -> QWidget:
        """Create a field selection widget."""
        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)

        # Field label
        label = QLabel(f"**{field_name}:**")
        label.setStyleSheet("font-weight: bold; font-size: 14px;")
        layout.addWidget(label)

        # Options from each scraper
        options_layout = QVBoxLayout()
        options_layout.setSpacing(5)

        combo = QComboBox()
        combo.addItem("-- Select Source --", None)

        for scraper_name in self.scrapers:
            scraper_data = self.sku_data["scrapers"][scraper_name]
            field_value = scraper_data.get(field_name, "")

            if field_value:
                # Format display: "SiteName: value"
                display_value = str(field_value)[:50]  # Truncate long values
                if len(str(field_value)) > 50:
                    display_value += "..."

                display_text = f"{scraper_name}: {display_value}"
                combo.addItem(display_text, scraper_name)

        combo.currentIndexChanged.connect(lambda: self.on_selection_changed(field_name, combo))
        options_layout.addWidget(combo)

        # Pre-select if we have a previous selection
        if field_name in self.selections:
            saved_scraper = self.selections[field_name]
            for i in range(combo.count()):
                if combo.itemData(i) == saved_scraper:
                    combo.setCurrentIndex(i)
                    break

        layout.addLayout(options_layout)

        self.field_widgets[field_name] = combo

        return container

    def _load_previous_selections(self):
        """Load previously saved selections for this product if they exist."""
        if self.results_hub:
            # Find if this product was already consolidated
            existing = next(
                (p for p in self.results_hub.consolidated_products if p["sku"] == self.sku), None
            )
            if existing and "fields" in existing:
                # Restore the selections
                for field_name, field_data in existing["fields"].items():
                    if "source" in field_data:
                        self.selections[field_name] = field_data["source"]

    def on_selection_changed(self, field_name: str, combo: QComboBox):
        """Handle field selection change."""
        selected_scraper = combo.currentData()

        if selected_scraper:
            self.selections[field_name] = selected_scraper
        elif field_name in self.selections:
            del self.selections[field_name]

    def on_image_source_changed(self):
        """Handle image source selection and load thumbnails."""
        selected_scraper = self.image_source_combo.currentData()

        if selected_scraper and selected_scraper in self.image_sources:
            self.selections["Images"] = selected_scraper
            images = self.image_sources[selected_scraper]

            # Clear existing thumbnails
            while self.thumbnail_layout.count() > 1:  # Keep the stretch
                item = self.thumbnail_layout.takeAt(0)
                if item.widget():
                    item.widget().deleteLater()

            # Load and display thumbnails
            for idx, img in enumerate(images):
                img_url = img if isinstance(img, str) else img.get("url", "")
                if img_url:
                    # Container for thumbnail + controls
                    thumb_container = QWidget()
                    thumb_layout = QVBoxLayout(thumb_container)
                    thumb_layout.setContentsMargins(0, 0, 0, 0)
                    thumb_layout.setSpacing(2)

                    # Thumbnail button
                    thumbnail_btn = QPushButton()
                    thumbnail_btn.setFixedSize(80, 80)
                    thumbnail_btn.setStyleSheet("""
                        QPushButton {
                            border: 2px solid #555;
                            padding: 2px;
                        }
                        QPushButton:hover {
                            border: 2px solid #888;
                        }
                    """)
                    thumbnail_btn.clicked.connect(
                        lambda checked, url=img_url: self.load_main_image(url)
                    )

                    # Load thumbnail
                    try:
                        import urllib.request

                        data = urllib.request.urlopen(img_url, timeout=5).read()
                        pixmap = QPixmap()
                        pixmap.loadFromData(data)
                        if not pixmap.isNull():
                            icon = QIcon(
                                pixmap.scaled(
                                    76,
                                    76,
                                    Qt.AspectRatioMode.KeepAspectRatio,
                                    Qt.TransformationMode.SmoothTransformation,
                                )
                            )
                            thumbnail_btn.setIcon(icon)
                            thumbnail_btn.setIconSize(QSize(76, 76))
                        else:
                            thumbnail_btn.setText(f"#{idx + 1}")
                    except:
                        thumbnail_btn.setText(f"#{idx + 1}")

                    thumb_layout.addWidget(thumbnail_btn)

                    # Reorder controls
                    control_layout = QHBoxLayout()
                    control_layout.setSpacing(2)

                    up_btn = QPushButton("▲")
                    up_btn.setFixedSize(38, 18)
                    up_btn.setStyleSheet("font-size: 10px; padding: 0;")
                    up_btn.clicked.connect(lambda checked, i=idx: self.move_image_up(i))
                    if idx == 0:
                        up_btn.setEnabled(False)
                    control_layout.addWidget(up_btn)

                    down_btn = QPushButton("▼")
                    down_btn.setFixedSize(38, 18)
                    down_btn.setStyleSheet("font-size: 10px; padding: 0;")
                    down_btn.clicked.connect(lambda checked, i=idx: self.move_image_down(i))
                    if idx == len(images) - 1:
                        down_btn.setEnabled(False)
                    control_layout.addWidget(down_btn)

                    thumb_layout.addLayout(control_layout)

                    self.thumbnail_layout.insertWidget(
                        self.thumbnail_layout.count() - 1, thumb_container
                    )

            # Load first image as main
            if images:
                first_url = images[0] if isinstance(images[0], str) else images[0].get("url", "")
                self.load_main_image(first_url)
        else:
            if "Images" in self.selections:
                del self.selections["Images"]
            self.main_image_label.clear()
            self.main_image_label.setText("Select an image source to preview")

            # Clear thumbnails
            while self.thumbnail_layout.count() > 1:
                item = self.thumbnail_layout.takeAt(0)
                if item.widget():
                    item.widget().deleteLater()

    def load_main_image(self, url: str):
        """Load an image into the main preview area."""
        try:
            import urllib.request

            data = urllib.request.urlopen(url, timeout=5).read()
            pixmap = QPixmap()
            pixmap.loadFromData(data)
            if not pixmap.isNull():
                scaled = pixmap.scaled(
                    480,
                    480,
                    Qt.AspectRatioMode.KeepAspectRatio,
                    Qt.TransformationMode.SmoothTransformation,
                )
                self.main_image_label.setPixmap(scaled)
            else:
                self.main_image_label.setText("Failed to load image")
        except Exception as e:
            self.main_image_label.setText(f"Error loading image:\\n{str(e)[:50]}")

    def move_image_up(self, index: int):
        """Move an image up in the order."""
        if index > 0:
            selected_scraper = self.image_source_combo.currentData()
            if selected_scraper and selected_scraper in self.image_sources:
                images = self.image_sources[selected_scraper]
                # Swap with previous
                images[index], images[index - 1] = images[index - 1], images[index]
                # Refresh thumbnails
                self.on_image_source_changed()

    def move_image_down(self, index: int):
        """Move an image down in the order."""
        selected_scraper = self.image_source_combo.currentData()
        if selected_scraper and selected_scraper in self.image_sources:
            images = self.image_sources[selected_scraper]
            if index < len(images) - 1:
                # Swap with next
                images[index], images[index + 1] = images[index + 1], images[index]
                # Refresh thumbnails
                self.on_image_source_changed()

    def get_consolidated_data(self) -> dict:
        """Get the consolidated product data."""
        consolidated = {"sku": self.sku, "fields": {}}

        for field, scraper in self.selections.items():
            value = self.sku_data["scrapers"][scraper].get(field)
            consolidated["fields"][field] = {"source": scraper, "value": value}

        return consolidated

    def save_consolidated(self):
        """Save the consolidated product."""
        if not self.selections:
            QMessageBox.warning(
                self, "No Selections", "Please select at least one field before saving."
            )
            return

        # Call ResultsHub's callback directly
        if self.results_hub:
            self.results_hub.on_consolidation_saved(self.get_consolidated_data())

    def save_and_close(self):
        """Save and exit consolidation."""
        self.save_consolidated()
        self.reject()

    def reject(self):
        """Cancel consolidation."""
        if self.results_hub:
            self.results_hub.on_consolidation_cancelled()

    def go_to_next(self):
        """Save current and move to next product."""
        # Auto-save current product
        if self.selections:
            self.save_consolidated()

        # Move to next
        if self.results_hub and self.current_index < len(self.consolidation_queue) - 1:
            self.results_hub.consolidation_index = self.current_index + 1
            next_sku = self.consolidation_queue[self.current_index + 1]["sku"]
            self.results_hub.consolidate_product(next_sku)
        else:
            # Last one - go back to list
            self.reject()

    def go_to_previous(self):
        """Move to previous product (auto-saving current)."""
        # Auto-save current product
        if self.selections:
            self.save_consolidated()

        if self.results_hub and self.current_index > 0:
            self.results_hub.consolidation_index = self.current_index - 1
            prev_sku = self.consolidation_queue[self.current_index - 1]["sku"]
            self.results_hub.consolidate_product(prev_sku)


class ResultsHub(QWidget):
    def __init__(self):
        super().__init__()

        # State variables
        self.current_session_file = None
        self.session_data = {}
        self.consolidated_data = []
        self.scrapers_in_session = []
        self.consolidated_products = []  # Manually consolidated products

        # Consolidation workflow state
        self.consolidation_queue = []  # List of products to consolidate
        self.consolidation_index = 0  # Current position in queue

        # Filter state
        self.search_term = ""
        self.show_diff_only = False

        # Create UI
        self.create_widgets()

        # Load history
        self.load_session_history()

    def create_widgets(self):
        """Create the main UI components."""
        # Main Layout
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # Stacked widget for different views
        self.view_stack = QStackedWidget()
        main_layout.addWidget(self.view_stack)

        # View 1: Session Browser (default)
        self.browser_view = self.create_browser_view()
        self.view_stack.addWidget(self.browser_view)

        # View 2: Consolidation View
        self.consolidation_container = QWidget()
        self.view_stack.addWidget(self.consolidation_container)

        # Set default view
        self.view_stack.setCurrentIndex(0)

    def create_browser_view(self) -> QWidget:
        """Create the session browser view."""
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # Splitter
        splitter = QSplitter(Qt.Orientation.Horizontal)
        layout.addWidget(splitter)

        # --- Left Sidebar: Session History ---
        sidebar_widget = QWidget()
        sidebar_layout = QVBoxLayout(sidebar_widget)
        sidebar_layout.setContentsMargins(10, 10, 10, 10)
        sidebar_layout.setSpacing(10)

        sidebar_label = QLabel("🕒 Scrape History")
        sidebar_label.setStyleSheet("font-weight: bold; font-size: 14px;")
        sidebar_layout.addWidget(sidebar_label)

        self.history_list = QListWidget()
        self.history_list.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.history_list.itemClicked.connect(self.on_session_selected)
        sidebar_layout.addWidget(self.history_list)

        refresh_btn = QPushButton("🔄 Refresh List")
        refresh_btn.clicked.connect(self.load_session_history)
        sidebar_layout.addWidget(refresh_btn)

        delete_btn = QPushButton("🗑️ Delete Session")
        delete_btn.setProperty("class", "danger")
        delete_btn.clicked.connect(self.delete_selected_session)
        sidebar_layout.addWidget(delete_btn)

        splitter.addWidget(sidebar_widget)

        # --- Right Content: Results Table ---
        content_widget = QWidget()
        content_layout = QVBoxLayout(content_widget)
        content_layout.setContentsMargins(15, 15, 15, 15)
        content_layout.setSpacing(10)

        # Header / Actions
        header_layout = QHBoxLayout()

        self.session_title = QLabel("Select a session to view results")
        self.session_title.setStyleSheet("font-size: 16px; font-weight: bold;")
        header_layout.addWidget(self.session_title)

        header_layout.addStretch()

        self.consolidate_btn = QPushButton("🔀 Consolidate Products")
        self.consolidate_btn.setProperty("class", "primary")
        self.consolidate_btn.clicked.connect(self.enter_consolidation_mode)
        self.consolidate_btn.setEnabled(False)
        header_layout.addWidget(self.consolidate_btn)

        self.export_excel_btn = QPushButton("📊 Export to Excel")
        self.export_excel_btn.setProperty("class", "secondary")
        self.export_excel_btn.clicked.connect(self.export_to_excel)
        self.export_excel_btn.setEnabled(False)
        header_layout.addWidget(self.export_excel_btn)

        self.import_db_btn = QPushButton("📥 Import to DB")
        self.import_db_btn.setProperty("class", "success")
        self.import_db_btn.clicked.connect(self.import_to_database)
        self.import_db_btn.setEnabled(False)
        header_layout.addWidget(self.import_db_btn)

        content_layout.addLayout(header_layout)

        # Filters
        filter_layout = QHBoxLayout()

        filter_layout.addWidget(QLabel("Search:"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search SKU or Name...")
        self.search_input.textChanged.connect(self.on_search_change)
        filter_layout.addWidget(self.search_input)

        self.diff_only_cb = QCheckBox("Show Multi-Source Products Only")
        self.diff_only_cb.stateChanged.connect(self.on_filter_change)
        filter_layout.addWidget(self.diff_only_cb)

        content_layout.addLayout(filter_layout)

        # Table
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(
            QAbstractItemView.EditTrigger.DoubleClicked
            | QAbstractItemView.EditTrigger.EditKeyPressed
        )
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.cellChanged.connect(self.on_cell_edited)
        self.table.cellDoubleClicked.connect(self.on_product_double_clicked)

        # Install custom delegate for better text editing on Name and Brand columns
        text_delegate = LargeTextDelegate()
        self.table.setItemDelegateForColumn(3, text_delegate)  # Name column
        self.table.setItemDelegateForColumn(4, text_delegate)  # Brand column

        content_layout.addWidget(self.table)

        # Status Bar
        self.status_label = QLabel("Ready")
        self.status_label.setStyleSheet("color: #888;")
        content_layout.addWidget(self.status_label)

        splitter.addWidget(content_widget)

        # Set initial splitter sizes (narrower sidebar for more content space)
        splitter.setSizes([175, 900])

        return widget

    def load_session_history(self):
        """Load list of JSON files from scraper_results directory."""
        self.history_list.clear()
        results_dir = Path(PROJECT_ROOT) / "data" / "scraper_results"

        if not results_dir.exists():
            results_dir.mkdir(parents=True, exist_ok=True)

        files = sorted(
            results_dir.glob("scrape_session_*.json"), key=os.path.getmtime, reverse=True
        )

        for file_path in files:
            # Parse filename for display
            try:
                timestamp_str = file_path.stem.replace("scrape_session_", "")
                # Format: YYYYMMDD_HHMMSS -> YYYY-MM-DD HH:MM
                display_text = f"{timestamp_str[:4]}-{timestamp_str[4:6]}-{timestamp_str[6:8]} {timestamp_str[9:11]}:{timestamp_str[11:13]}"
            except:
                display_text = file_path.name

            item = QListWidgetItem(f"📄 {display_text}")
            item.setData(Qt.ItemDataRole.UserRole, str(file_path))
            self.history_list.addItem(item)

    def on_session_selected(self, item):
        """Handle session selection from sidebar."""
        file_path = item.data(Qt.ItemDataRole.UserRole)
        self.load_session_file(file_path)

    def load_session_file(self, file_path):
        """Load and parse a session JSON file."""
        try:
            with open(file_path, encoding="utf-8") as f:
                data = json.load(f)

            self.current_session_file = file_path
            self.session_data = data
            self.session_title.setText(f"Session: {os.path.basename(file_path)}")

            # Parse results
            self.parse_session_data()

            # Update UI
            self.update_table()
            self.consolidate_btn.setEnabled(True)
            self.import_db_btn.setEnabled(True)
            self.export_excel_btn.setEnabled(True)
            self.status_label.setText(f"Loaded {len(self.consolidated_data)} unique SKUs")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load session: {e}")

    def parse_session_data(self):
        """Parse session data into consolidated structure."""
        self.consolidated_data = []
        self.scrapers_in_session = []

        if not self.session_data:
            return

        raw_results = {}

        # Format: New Session Format
        if "session_id" in self.session_data and "results" in self.session_data:
            raw_results = self.session_data["results"]  # {scraper: {sku: data}}
            self.scrapers_in_session = list(raw_results.keys())
        elif isinstance(self.session_data, list):
            self.scrapers_in_session = ["Unknown"]
            raw_results = {
                "Unknown": {item.get("SKU", "Unknown"): item for item in self.session_data}
            }

        # Pivot data by SKU
        sku_map = {}

        # Extract price metadata if available
        price_data = self.session_data.get("metadata", {}).get("price", {})

        for scraper_name, scraper_data in raw_results.items():
            for sku, item_data in scraper_data.items():
                # Extract actual data payload
                payload = item_data.get("data", item_data)

                if sku not in sku_map:
                    sku_map[sku] = {
                        "sku": sku,
                        "price": price_data.get(sku, ""),  # Attach preserved Price
                        "scrapers": {},
                    }

                # Store full scraper data
                sku_map[sku]["scrapers"][scraper_name] = payload

        self.consolidated_data = list(sku_map.values())

    def on_search_change(self):
        self.search_term = self.search_input.text().strip().lower()
        self.update_table()

    def on_filter_change(self):
        self.show_diff_only = self.diff_only_cb.isChecked()
        self.update_table()

    def update_table(self):
        """Render the consolidated data to the table."""
        if not self.consolidated_data:
            self.table.setRowCount(0)
            return

        # Define columns
        columns = ["SKU", "Price", "Images", "Name", "Brand", "Status", "Actions"]

        self.table.setColumnCount(len(columns))
        self.table.setHorizontalHeaderLabels(columns)

        # Set column resize modes
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)  # SKU
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)  # Price
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Interactive)  # Images - interactive with max width
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)  # Name - expands
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)  # Brand
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)  # Status
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.Interactive)  # Actions - interactive with fixed width

        # Set reasonable widths for interactive columns
        header.resizeSection(2, 200)  # Images column default 200px
        header.resizeSection(6, 100)  # Actions column default 100px

        # Set maximum widths for interactive columns
        header.setMaximumSectionSize(300)  # Global max for all sections

        # Filter data
        filtered_data = []
        for item in self.consolidated_data:
            # Search Filter
            if self.search_term:
                search_text = f"{item['sku']}".lower()
                for scraper_data in item["scrapers"].values():
                    search_text += (
                        f" {scraper_data.get('Name', '')} {scraper_data.get('Brand', '')}".lower()
                    )

                if self.search_term not in search_text:
                    continue

            # Multi-source filter
            if self.show_diff_only and len(item["scrapers"]) < 2:
                continue

            filtered_data.append(item)

        # Populate rows
        self.table.setRowCount(len(filtered_data))

        for row_idx, item in enumerate(filtered_data):
            sku = item["sku"]
            # sources = ", ".join(item["scrapers"].keys()) # Removed

            # Get name and brand - only show if single source, otherwise placeholder
            num_sources = len(item["scrapers"])
            if num_sources == 1:
                # Single source - show the values
                scraper_data = next(iter(item["scrapers"].values()))
                name = scraper_data.get("Name", "")
                brand = scraper_data.get("Brand", "")
            else:
                # Multiple sources - show placeholders
                name = f"Multiple sources ({num_sources})"
                brand = f"Multiple sources ({num_sources})"

            # Get price
            price = item.get("price", "")

            # Get images string
            if num_sources == 1:
                # Single source - show the images
                scraper_data = next(iter(item["scrapers"].values()))
                imgs = scraper_data.get("Images", [])
                images_str = ", ".join(
                    [img if isinstance(img, str) else img.get("url", "") for img in imgs]
                ) if isinstance(imgs, list) and imgs else ""
            else:
                # Multiple sources - show placeholder
                images_str = f"Multiple sources ({num_sources})"

            # Determine status
            status = "Consolidated" if sku in [p["sku"] for p in self.consolidated_products] else "Pending"

            # Create action button
            action_btn = QPushButton("Consolidate" if status == "Pending" else "View")
            action_btn.clicked.connect(lambda checked, s=sku: self.consolidate_product(s))
            action_widget = QWidget()
            action_layout = QHBoxLayout(action_widget)
            action_layout.addWidget(action_btn)
            action_layout.setContentsMargins(0, 0, 0, 0)

            # Set table items
            self.table.setItem(row_idx, 0, QTableWidgetItem(sku))
            self.table.setItem(row_idx, 1, QTableWidgetItem(price))
            self.table.setItem(row_idx, 2, QTableWidgetItem(images_str))
            self.table.setItem(row_idx, 3, QTableWidgetItem(name))
            self.table.setItem(row_idx, 4, QTableWidgetItem(brand))
            self.table.setItem(row_idx, 5, QTableWidgetItem(status))
            self.table.setCellWidget(row_idx, 6, action_widget)

    def on_product_double_clicked(self, row, col):
        """Handle double-click on product row - only trigger consolidation on SKU or Status columns."""
        # Allow editing for Name (col 3) and Brand (col 4)
        if col in [3, 4]:
            return  # Let the edit happen

        # For other columns (SKU, Sources, Status), trigger consolidation
        sku_item = self.table.item(row, 0)
        if sku_item:
            sku = sku_item.text()
            self.consolidate_product(sku)

    def on_cell_edited(self, row, col):
        """Handle cell editing for Name and Brand columns."""
        if col not in [3, 4]:  # Only Name and Brand are editable
            return

        sku_item = self.table.item(row, 0)
        if not sku_item:
            return
        sku = sku_item.text()

        edited_item = self.table.item(row, col)
        if not edited_item:
            return
        new_value = edited_item.text()

        # Find the product in consolidated_data
        product = next((p for p in self.consolidated_data if p["sku"] == sku), None)
        if not product:
            return

        field_name = "Name" if col == 3 else "Brand"

        # Update all scraper data for this product
        for scraper_data in product["scrapers"].values():
            scraper_data[field_name] = new_value

        # Also update consolidated_products if it exists
        for cons_prod in self.consolidated_products:
            if cons_prod["sku"] == sku:
                if field_name in cons_prod["fields"]:
                    cons_prod["fields"][field_name]["value"] = new_value
                break

    def enter_consolidation_mode(self):
        """Enter consolidation mode - iterate through ALL unconsolidated products."""
        # Get all products that aren't consolidated yet
        products_to_consolidate = [
            item
            for item in self.consolidated_data
            if item["sku"] not in [p["sku"] for p in self.consolidated_products]
        ]

        if not products_to_consolidate:
            QMessageBox.information(
                self, "All Done", "All products have been verified!"
            )
            return

        # Set up consolidation queue
        self.consolidation_queue = products_to_consolidate
        self.consolidation_index = 0

        # Start with first product (no popup)
        self.consolidate_product(products_to_consolidate[0]["sku"])

    def consolidate_product(self, sku: str, force: bool = False):
        """Open consolidation view for a specific product."""
        # Find product data
        product_data = next((p for p in self.consolidated_data if p["sku"] == sku), None)

        if not product_data:
            return

        # Removed single-source check to allow verification of all products

        # Create consolidation widget with ResultsHub reference
        consolidation_widget = ConsolidationWidget(product_data, results_hub_parent=self)

        # Get or create consolidation container layout
        layout = self.consolidation_container.layout()
        if layout is None:
            layout = QVBoxLayout(self.consolidation_container)

        # Clear existing widgets
        while layout.count():
            item = layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        # Add back button
        back_btn = QPushButton("← Back to Results")
        back_btn.clicked.connect(self.exit_consolidation_mode)
        layout.addWidget(back_btn)

        layout.addWidget(consolidation_widget)

        # Switch to consolidation view
        self.view_stack.setCurrentIndex(1)

    def exit_consolidation_mode(self):
        """Return to browser view."""
        self.view_stack.setCurrentIndex(0)

    def on_consolidation_saved(self, consolidated_data: dict):
        """Handle saved consolidation."""
        # Check if this product was already consolidated (update instead of append)
        existing_idx = next(
            (
                i
                for i, p in enumerate(self.consolidated_products)
                if p["sku"] == consolidated_data["sku"]
            ),
            None,
        )

        if existing_idx is not None:
            # Update existing
            self.consolidated_products[existing_idx] = consolidated_data
        else:
            # Add new
            self.consolidated_products.append(consolidated_data)

        # Return to browser and refresh (no popup)
        self.exit_consolidation_mode()
        self.update_table()

    def on_consolidation_cancelled(self):
        """Handle cancelled consolidation."""
        self.exit_consolidation_mode()

    def import_to_database(self):
        """Import consolidated products to the database."""
        if not self.consolidated_products:
            QMessageBox.warning(
                self, "No Data", "Please consolidate some products first before importing."
            )
            return

        reply = QMessageBox.question(
            self,
            "Confirm Import",
            f"Import {len(self.consolidated_products)} consolidated products into the database?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                self._perform_database_import()
                QMessageBox.information(
                    self,
                    "Success",
                    f"Successfully imported {len(self.consolidated_products)} products!",
                )
            except Exception as e:
                QMessageBox.critical(self, "Import Failed", f"Error: {e}")

    def _perform_database_import(self):
        """Perform the actual database import."""
        if not DB_PATH.parent.exists():
            DB_PATH.parent.mkdir(parents=True, exist_ok=True)

        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        # Create table if not exists
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS products (
                SKU TEXT PRIMARY KEY,
                Name TEXT,
                Brand TEXT,
                Price TEXT,
                Weight TEXT,
                Category TEXT,
                Product_Type TEXT,
                Images TEXT,
                last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        for product in self.consolidated_products:
            sku = product["sku"]
            fields = product["fields"]

            # Extract values
            name = fields.get("Name", {}).get("value", "")
            brand = fields.get("Brand", {}).get("value", "")
            price = fields.get("Price", {}).get("value", "")
            weight = fields.get("Weight", {}).get("value", "")
            category = fields.get("Category", {}).get("value", "")
            product_type = fields.get("Product_Type", {}).get("value", "")
            images = str(fields.get("Images", {}).get("value", ""))

            # Upsert
            cursor.execute(
                """
                INSERT OR REPLACE INTO products 
                (SKU, Name, Brand, Price, Weight, Category, Product_Type, Images, last_updated)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            """,
                (sku, name, brand, price, weight, category, product_type, images),
            )

        conn.commit()
        conn.close()

    def export_to_excel(self):
        """Export products to ShopSite-compatible Excel file."""
        if not self.consolidated_data:
            QMessageBox.warning(self, "No Data", "No session data loaded to export.")
            return

        default_filename = f"products_export_{self.get_timestamp()}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export to Excel", default_filename, "Excel Files (*.xlsx)"
        )

        if not file_path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Products"
            headers = ["SKU", "Name", "Brand", "Weight", "Image URLs", "Price"]
            ws.append(headers)

            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font

            for item in self.consolidated_data:
                sku = item["sku"]
                consolidated = next(
                    (p for p in self.consolidated_products if p["sku"] == sku), None
                )

                if consolidated:
                    fields = consolidated["fields"]
                    name = fields.get("Name", {}).get("value", "")
                    brand = fields.get("Brand", {}).get("value", "")
                    weight = fields.get("Weight", {}).get("value", "")
                    images_data = fields.get("Images", {}).get("value", [])
                    image_urls = (
                        ", ".join(
                            [
                                img if isinstance(img, str) else img.get("url", "")
                                for img in images_data
                            ]
                        )
                        if isinstance(images_data, list)
                        else str(images_data or "")
                    )
                    price = item.get("price", "")  # Get preserved Price from metadata
                else:
                    name, brand, weight, image_urls, price = "", "", "", "", item.get("price", "")
                    for scraper_data in item["scrapers"].values():
                        if not name:
                            name = scraper_data.get("Name", "")
                        if not brand:
                            brand = scraper_data.get("Brand", "")
                        if not weight:
                            weight = scraper_data.get("Weight", "")
                        if not image_urls:
                            imgs = scraper_data.get("Images", [])
                            if isinstance(imgs, list) and imgs:
                                image_urls = ", ".join(
                                    [
                                        img if isinstance(img, str) else img.get("url", "")
                                        for img in imgs
                                    ]
                                )
                        if name and brand and weight and image_urls:
                            break
                ws.append([sku, name, brand, weight, image_urls, price])

            for column in ws.columns:
                max_length = max(len(str(cell.value)) for cell in column)
                ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

            wb.save(file_path)
            QMessageBox.information(
                self,
                "Success",
                f"Exported {len(self.consolidated_data)} products to:\\n{file_path}",
            )
        except ImportError:
            QMessageBox.critical(
                self,
                "Missing Dependency",
                "openpyxl library is required.\\nInstall with: pip install openpyxl",
            )
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Error: {e}")

    def get_timestamp(self):
        """Get current timestamp for filename."""
        from datetime import datetime

        return datetime.now().strftime("%Y%m%d_%H%M%S")

    def delete_selected_session(self):
        """Delete the currently selected session file."""
        selected_item = self.history_list.currentItem()
        if not selected_item:
            QMessageBox.warning(self, "No Selection", "Please select a session to delete.")
            return

        file_path = selected_item.data(Qt.ItemDataRole.UserRole)
        session_name = selected_item.text()

        # Confirmation dialog
        reply = QMessageBox.question(
            self,
            "Confirm Delete",
            f"Are you sure you want to delete this session?\n\n{session_name}\n\nThis action cannot be undone.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                import os

                os.remove(file_path)

                # Clear current session if it was the deleted one
                if self.current_session_file == file_path:
                    self.current_session_file = None
                    self.session_data = {}
                    self.consolidated_data = []
                    self.session_title.setText("Select a session to view results")
                    self.table.setRowCount(0)
                    self.consolidate_btn.setEnabled(False)
                    self.import_db_btn.setEnabled(False)
                    self.export_excel_btn.setEnabled(False)

                # Refresh the list
                self.load_session_history()
                QMessageBox.information(self, "Success", "Session deleted successfully.")

            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to delete session: {e}")
